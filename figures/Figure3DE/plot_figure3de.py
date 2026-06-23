"""
Figure 3D/E 그래프 생성 스크립트
rates 파일 완료 후 실행

사용법:
    cd ~/cre_reproduction/neoParSA/tests/transcpp/fits
    python3 /path/to/plot_figure3de.py

출력:
    Figure3DE/output/Figure3D.png
    Figure3DE/output/Figure3E.png
    Figure3DE/output/Figure3DE.png
"""

import os, re, glob
import numpy as np
import openpyxl
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

BASE   = os.path.expanduser("~/Desktop/SB_final/ScienceDirect_files_26May2026_04-41-20.049")
FITS   = os.path.expanduser("~/cre_reproduction/neoParSA/tests/transcpp/fits")
MMC4   = f"{BASE}/1-s2.0-S2589004223028249-mmc4.xlsx"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT, exist_ok=True)

COLORS = {1:'#9DB8DE',2:'#E87513',3:'#4169A1',4:'#70AD47',
          5:'#00A6C6',6:'#8C8C8C',7:'#8064A2'}

def load_data():
    wb      = openpyxl.load_workbook(MMC4)
    rows_s  = list(wb['single-hit'].iter_rows(values_only=True))[1:]
    wt_s    = next(r[2] for r in rows_s if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
    meas    = {r[0]: np.log2(r[2]/wt_s) for r in rows_s
               if r[0] and r[2] and re.search(r'pos_(\d+)_([ATGC])$', str(r[0]))}
    rows_m  = list(wb['multi-hit'].iter_rows(values_only=True))[1:]
    wt_m    = next(r[2] for r in rows_m if 'WT' in str(r[0]))
    return meas, rows_m, wt_m

def calc_r_single(rates_file, meas_dict):
    try:
        with open(rates_file) as f: lines = f.readlines()
        if len(lines) < 2: return None
        header = lines[0].strip().split()
        data   = lines[1].strip().split()
        pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
        wt_p   = pred.get('synCRE_Promega_0')
        if not wt_p or wt_p <= 0: return None
        pairs  = [(mv, np.log2(pred[n]/wt_p)) for n,mv in meas_dict.items() if n in pred and pred[n]>0]
        if len(pairs) < 10: return None
        r, _ = stats.pearsonr(*zip(*pairs))
        return r if not np.isnan(r) else None
    except: return None

def calc_r_multihit(rates_file, rows_m, wt_m):
    try:
        with open(rates_file) as f: lines = f.readlines()
        if len(lines) < 2: return None
        header = lines[0].strip().split()
        data   = lines[1].strip().split()
        pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
        wt_p   = pred.get('synCRE_Promega_0_WT') or pred.get('synCRE_Promega_0')
        if not wt_p or wt_p <= 0: return None
        pairs  = [(np.log2(float(e)/float(wt_m)), np.log2(pred[n]/wt_p))
                  for n,seq,e in rows_m if 'WT' not in str(n) and n in pred and pred[n]>0]
        if len(pairs) < 10: return None
        r, _ = stats.pearsonr(*zip(*pairs))
        return r if not np.isnan(r) else None
    except: return None

def draw_boxplot(ax, results, title, label):
    n_list   = sorted(results.keys())
    data_box = [results[n] for n in n_list]
    if not n_list:
        ax.text(0.5, 0.5, 'No data', ha='center', va='center', transform=ax.transAxes)
        ax.set_title(title, fontsize=11, fontweight='bold')
        return

    ax.set_facecolor('white')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', color='#E0E0E0', linewidth=0.8, zorder=0)

    bp = ax.boxplot(data_box, patch_artist=True, widths=0.5,
                    medianprops=dict(color='white', linewidth=2.5),
                    whiskerprops=dict(linewidth=1.2, color='#888888'),
                    capprops=dict(linewidth=1.2, color='#888888'),
                    flierprops=dict(marker='o', markersize=4, markerfacecolor='#888888',
                                    markeredgecolor='none', alpha=0.7),
                    zorder=3)
    for patch, n in zip(bp['boxes'], n_list):
        patch.set_facecolor(COLORS.get(n,'steelblue'))
        patch.set_alpha(0.85); patch.set_edgecolor('none')

    for i, (n, vals) in enumerate(zip(n_list, data_box)):
        ax.scatter(i+1, np.mean(vals), s=60, color='black', zorder=5)

    ax.set_xticks(range(1, len(n_list)+1))
    ax.set_xticklabels([str(n) for n in n_list], fontsize=11)
    ax.set_xlabel('number of TFs', fontsize=12)
    ax.set_ylabel("Pearson's R", fontsize=12)
    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.set_ylim(-0.4, 1.0)
    ax.text(0.02, 0.98, label, transform=ax.transAxes, fontsize=14, fontweight='bold', va='top')

def main():
    print("=== Figure 3D/E 그래프 생성 ===")
    meas_dict, rows_m, wt_m = load_data()

    # Fig3D: single-hit rates
    results_d = {}
    for rf in glob.glob(f'{FITS}/fig3de_xmls/*.rates'):
        n = int(rf.split('fig3de_n')[-1].split('_')[0])
        r = calc_r_single(rf, meas_dict)
        if r is not None: results_d.setdefault(n, []).append(r)

    # Fig3E: multi-hit rates
    results_e = {}
    for rf in glob.glob(f'{FITS}/fig3de_mh_xmls/*_mh_v2.xml.rates'):
        n = int(rf.split('fig3de_n')[-1].split('_')[0])
        r = calc_r_multihit(rf, rows_m, wt_m)
        if r is not None: results_e.setdefault(n, []).append(r)

    print("Fig3D:", {n: len(v) for n,v in sorted(results_d.items())})
    print("Fig3E:", {n: len(v) for n,v in sorted(results_e.items())})

    # 합친 그래프
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    draw_boxplot(axes[0], results_d, '(D) With self-competition\nSingle-hit fitting', 'D')
    draw_boxplot(axes[1], results_e, '(E) With self-competition\nMulti-hit prediction', 'E')
    plt.tight_layout()
    plt.savefig(f"{OUTPUT}/Figure3DE.png", dpi=150, bbox_inches='tight')

    # 개별 저장
    for results, label, title, fname in [
        (results_d, 'D', '(D) With self-competition\nSingle-hit fitting', 'Figure3D.png'),
        (results_e, 'E', '(E) With self-competition\nMulti-hit prediction', 'Figure3E.png'),
    ]:
        fig, ax = plt.subplots(figsize=(8, 6))
        draw_boxplot(ax, results, title, label)
        plt.tight_layout()
        plt.savefig(f"{OUTPUT}/{fname}", dpi=150, bbox_inches='tight')
        plt.close()

    print(f"✅ 저장 완료: {OUTPUT}/")

if __name__ == '__main__':
    main()
