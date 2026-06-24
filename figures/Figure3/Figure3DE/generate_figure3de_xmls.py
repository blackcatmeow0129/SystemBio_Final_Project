"""
Figure 3D/E XML 생성 스크립트
127 TF 조합 × 8 모델 = 1016개 XML 생성 (SelfCompetition=true)

사용법:
    python3 generate_figure3de_xmls.py

출력:
    ~/cre_reproduction/.../fig3de_xmls/ 에 1016개 XML 파일
"""

import os, pickle, math, requests
import openpyxl
from itertools import combinations

BASE     = os.path.expanduser("~/Desktop/SB_final/ScienceDirect_files_26May2026_04-41-20.049")
FITS     = os.path.expanduser("~/cre_reproduction/neoParSA/tests/transcpp/fits")
MMC2     = f"{BASE}/1-s2.0-S2589004223028249-mmc2.xlsx"
MMC3     = f"{BASE}/1-s2.0-S2589004223028249-mmc3.xlsx"
MMC4     = f"{BASE}/1-s2.0-S2589004223028249-mmc4.xlsx"
XML_DIR  = f"{FITS}/fig3de_xmls"
os.makedirs(XML_DIR, exist_ok=True)

TFS = ["CREB1","CREB3","CREB5","CREM","ATF1","ATF4","ATF7"]

def load_pwms():
    ws = openpyxl.load_workbook(MMC3)['시트1']
    pwms = {}
    skip = {"Source: Cis-BP","Source: Weirauch14","Source: Jolma_367",
            "Source: Jolma_371","Source: Jolma_353","Source: Jolma_354",
            "Source: Badis09, Zhao and Stormo 11"}
    tf, cur = None, {"A":[],"C":[],"G":[],"T":[]}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[0] not in skip:
            if tf: pwms[tf] = cur
            tf, cur = row[0], {"A":[],"C":[],"G":[],"T":[]}
        if row[1] in "ACGT":
            cur[row[1]] = [v for v in row[2:] if v is not None]
    if tf: pwms[tf] = cur
    return pwms

def load_params():
    ws    = openpyxl.load_workbook(MMC2)['7TF_self']
    rows  = list(ws.iter_rows(values_only=True))
    mrow  = next(r for r in rows if r[0]=='TF' and r[1]=='Model num')
    mcols = {int(mrow[c]):c for c in range(2,10)}
    params = {m:{} for m in range(1,9)}
    tf = None
    for row in rows:
        if row[0]=='TF': continue
        if row[0] is not None: tf = row[0]
        if row[1] in ['Rmax','Theta']:
            for m,c in mcols.items(): params[m][row[1]] = row[c]
        elif row[1] in ['coef','kmax','lambda','threshold'] and tf:
            for m,c in mcols.items(): params[m].setdefault(tf,{})[row[1]] = row[c]
    return params

def load_single_hit_data():
    rows = list(openpyxl.load_workbook(MMC4)['single-hit'].iter_rows(values_only=True))[1:]
    gene_lines = "\n".join([f'        <Gene name="{r[0]}" header="{r[0]}" left_bound="-235" right_bound="-38" TSS="-1" promoter="basic"/>' for r in rows if r[0]])
    gene_cols  = " ".join([f'{r[0]}="{r[2]:.4f}"' for r in rows if r[0] and r[2]])
    rate_line  = f'      <TableRow ID="theonlyone" {gene_cols}/>'
    return gene_lines, rate_line

def make_pwm_xml(d):
    lines = ['        <PWM type="PSSM">','          <base>A C G T</base>']
    for i in range(len(d['A'])):
        lines.append(f'          <position>{d["A"][i]:10.6f}; {d["C"][i]:10.6f}; {d["G"][i]:10.6f}; {d["T"][i]:10.6f}</position>')
    lines.append('        </PWM>')
    return "\n".join(lines)

def make_tf_xml(name, p, pwm):
    DEFAULT = {"coef":1.0,"kmax":1.0,"lambda":2.0,"threshold":0.0}
    p = {**DEFAULT, **p}
    return f"""      <TF name="{name}" bsize="14" include="true">
        <kmax      value="{p['kmax']}"      lim_low="1e-06" lim_high="4.0"  anneal="true" move="Kmax"/>
        <threshold value="{p['threshold']}" lim_low="-5"   lim_high="15.0" anneal="true" move="Sites"/>
        <lambda    value="{p['lambda']}"    lim_low="0.5"  lim_high="5.0"  anneal="true" move="Lambda"/>
        <Coefficients>
          <coef value="{p['coef']}" lim_low="0.0001" lim_high="20.0" anneal="true" move="Promoter"/>
        </Coefficients>
{make_pwm_xml(pwm)}
      </TF>"""

def make_xml(tfs_xml, tf_data, gene_lines, rate_line, theta, seed, self_comp=True):
    sc = "true" if self_comp else "false"
    return f"""<?xml version="1.0" encoding="utf-8"?>
<Root>
  <annealer_input init_T="100000" lambda="0.0001" init_loop="100000"/>
  <move interval="100" gain="3"/>
  <count_criterion freeze_crit="10" freeze_cnt="5"/>
  <mix adaptcoef="10"/>
  <lam tau="100" memLength_mean=".200" memLength_sd="100" criterion="10" freeze_cnt="5"/>
  <Mode>
    <Verbose value="1"/>
    <ScoreFunction value="sse"/>
    <GCcontent value="0.534"/>
    <NumThreads value="1"/>
    <SelfCompetition value="{sc}"/>
    <Precision value="8"/>
    <Seed value="{seed}"/>
  </Mode>
  <Input>
    <Distances>
      <Distance name="Quenching" distfunc="Trapezoid">
        <A value="100" lim_low="50" lim_high="100" anneal="false" move="Quenching"/>
        <B value="50"  lim_low="50" lim_high="50"  anneal="false" move="Quenching"/>
      </Distance>
    </Distances>
    <Promoters>
      <Promoter name="basic" function="Arrhenius2">
        <Q     value="1"       lim_low="1"   lim_high="1"   anneal="false" move="Promoter"/>
        <Rmax  value="200"     lim_low="200" lim_high="200" anneal="false" move="Promoter"/>
        <Theta value="{theta}" lim_low="5"   lim_high="20"  anneal="true"  move="Promoter"/>
      </Promoter>
    </Promoters>
    <TFs>
{tfs_xml}
    </TFs>
    <Interactions/>
    <Genes>
      <Source name="MRPA" file="sequences/MRPA_baseM.fa" type="fasta">
{gene_lines}
      </Source>
    </Genes>
    <ScaleFactors>
      <ScaleFactor name="default">
        <A value="1" lim_low="1" lim_high="1" anneal="false"/>
        <B value="0" lim_low="0" lim_high="0" anneal="false"/>
      </ScaleFactor>
    </ScaleFactors>
    <RateData row="ID" col="gene">
{rate_line}
    </RateData>
    <TFData row="ID" col="TF">
      <TableRow ID="theonlyone" {tf_data}/>
    </TFData>
  </Input>
</Root>"""

def main():
    print("=== Figure 3D/E XML 생성 시작 ===")
    pwms        = load_pwms()
    params      = load_params()
    gene_lines, rate_line = load_single_hit_data()

    count = 0
    for n_tfs in range(1, 8):
        for combo in combinations(TFS, n_tfs):
            combo_name = "_".join(combo)
            for m_num in range(1, 9):
                mp       = params[m_num]
                theta    = mp.get('Theta', 6.0)
                tf_data  = " ".join([f'{tf}="100"' for tf in combo])
                tfs_xml  = "\n".join([make_tf_xml(tf, mp.get(tf,{}), pwms.get(tf, pwms['CREB1'])) for tf in combo])

                xml      = make_xml(tfs_xml, tf_data, gene_lines, rate_line, theta, m_num, self_comp=True)
                fname    = f"{XML_DIR}/fig3de_n{n_tfs}_m{m_num}_{combo_name}.xml"
                with open(fname, 'w') as f:
                    f.write(xml)
                count += 1

    print(f"✅ {count}개 XML 생성 완료! → {XML_DIR}")
    print("다음 단계: bash run_figure3de_local.sh (Mac) 또는 run_figure3de_server.sh (서버)")

if __name__ == '__main__':
    main()
