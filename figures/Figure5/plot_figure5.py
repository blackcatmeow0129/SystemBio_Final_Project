"""
Reproduce Figure 5 panels for A/T substitution functional binding-site analysis.

The script uses the supplementary workbooks in the parent directory:
  - mmc2: model parameters
  - mmc3: TF PWMs
  - mmc4: MPRA single-hit sequences/expression

If a transcpp .rates file is present, the lower activity panels use model
predictions. Otherwise those panels are left empty and the MPRA/TFBS/DDA panels
are still generated.
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import math
import os
import re
import subprocess
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

try:
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    import numpy as np
    import openpyxl
except ModuleNotFoundError as exc:
    raise SystemExit(
        f"Missing Python package: {exc.name}\n"
        "Install the plotting dependencies first:\n"
        "  pip install openpyxl numpy matplotlib"
    ) from exc


BASES = "ACGT"
RC = str.maketrans("ACGT", "TGCA")
TFS = ["ATF7", "ATF1", "CREM", "CREB1"]
TF_COLORS = {
    "ATF7": "#b83cff",
    "ATF1": "#30309b",
    "CREM": "#4caf23",
    "CREB1": "#ff9800",
}
CRE_SITES = {
    "CRE1": (8, 19, "#ff6b6b"),
    "CRE2": (35, 44, "#ff6b6b"),
    "CRE3": (44, 52, "#ff6b6b"),
    "cryptic": (60, 68, "#5568d9"),
    "CRE4": (67, 79, "#ff6b6b"),
}
CRE1_VARIANTS_A = [(11, "A"), (14, "A"), (17, "A")]
CRE4_VARIANTS_A = [(69, "A"), (72, "A"), (75, "A")]
CRE1_VARIANTS_T = [(12, "T"), (15, "T"), (18, "T")]
CRE4_VARIANTS_T = [(70, "T"), (73, "T"), (76, "T")]
LEFT_BOUND = -235


def label_to_index(pos_label: int) -> int:
    return pos_label


def variant_gene(pos_label: int, base: str) -> str:
    return f"scanmut_single_pos_{label_to_index(pos_label)}_{base}"


@dataclass
class TFParam:
    coef: float
    kmax: float
    lam: float
    threshold: float


@dataclass
class Site:
    tf: str
    start: int
    end: int
    score: float
    occupancy: float
    strand: str
    index: int = 0


def find_one(pattern: str, root: Path) -> Path:
    matches = sorted(root.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching {pattern} under {root}")
    return matches[0]


def load_mmc2_params(path: Path) -> dict[str, TFParam]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["4TF_self"]
    params: dict[str, dict[str, float]] = {tf: {} for tf in ["ATF1", "ATF7", "CREB1", "CREM"]}

    rows = list(ws.iter_rows(values_only=True))
    header = next(row for row in rows if row and row[0] == "TF" and row[1] == "Model num")
    model5_col = next(i for i, value in enumerate(header) if i >= 2 and value is not None and float(value) == 5.0)

    for row in rows:
        if not row or row[0] not in params or row[1] not in {"coef", "kmax", "lambda", "threshold"}:
            continue
        key = "lam" if row[1] == "lambda" else row[1]
        params[row[0]][key] = float(row[model5_col])

    return {tf: TFParam(**values) for tf, values in params.items()}


def load_xml_output_params(path: Path) -> dict[str, TFParam]:
    tree = ET.parse(path)
    root = tree.getroot()
    section = root.find("Output")
    if section is None:
        section = root.find("Input")
    if section is None:
        raise ValueError(f"No Output/Input section found in {path}")

    params: dict[str, TFParam] = {}
    for tf_node in section.findall("./TFs/TF"):
        tf = tf_node.attrib.get("name")
        if tf not in TFS:
            continue
        coef_node = tf_node.find("./Coefficients/coef")
        params[tf] = TFParam(
            coef=float(coef_node.attrib["value"]),
            kmax=float(tf_node.find("kmax").attrib["value"]),
            lam=float(tf_node.find("lambda").attrib["value"]),
            threshold=float(tf_node.find("threshold").attrib["value"]),
        )
    missing = [tf for tf in TFS if tf not in params]
    if missing:
        raise ValueError(f"Missing TF params in {path}: {missing}")
    return params


def load_pwm(path: Path) -> dict[str, dict[str, list[float]]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    skip_prefix = "Source:"
    pwm: dict[str, dict[str, list[float]]] = {}
    current_tf: str | None = None
    current_rows = {base: [] for base in BASES}

    for row in ws.iter_rows(values_only=True):
        label = row[0] if len(row) > 0 else None
        base = row[1] if len(row) > 1 else None

        if label and isinstance(label, str) and not label.startswith(skip_prefix):
            if current_tf:
                pwm[current_tf] = current_rows
            current_tf = label
            current_rows = {b: [] for b in BASES}

        if isinstance(base, str) and base in BASES:
            current_rows[base] = [float(v) for v in row[2:] if v is not None]

    if current_tf:
        pwm[current_tf] = current_rows
    return pwm


def load_single_hit(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = list(wb["single-hit"].iter_rows(values_only=True))[1:]
    wt_name, wt_seq, wt_expr = rows[0]
    variants = []
    active_start = None

    for name, seq, expr in rows[1:]:
        match = re.search(r"pos_(\d+)_([ACGT])$", name)
        if not match:
            continue
        pos = int(match.group(1))
        base = match.group(2)
        diffs = [idx for idx, (a, b) in enumerate(zip(wt_seq, seq)) if a != b]
        if diffs and active_start is None:
            active_start = diffs[0] - pos
        variants.append({"name": name, "pos": pos, "base": base, "seq": seq, "expr": float(expr)})

    if active_start is None:
        raise ValueError("Could not infer active CRE start from single-hit variants")

    active_len = max(v["pos"] for v in variants) + 1
    wt_cre = wt_seq[active_start : active_start + active_len]
    return wt_seq, wt_cre, float(wt_expr), active_start, variants


def load_rates(path: Path | None) -> dict[str, float]:
    if not path or not path.exists():
        return {}
    raw = path.read_bytes()
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        text = raw.decode("utf-16")
    else:
        text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    if len(lines) < 2:
        return {}
    header = lines[0].split()[1:]
    values = lines[1].split()[1:]
    return {name: float(value) for name, value in zip(header, values)}


def parse_unfold_occupancy(text: str) -> dict[str, float]:
    lines = [line for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        return {}
    headers = lines[0].split()[1:]
    values = lines[1].split()[1:]
    return {header: float(value) for header, value in zip(headers, values)}


def parse_unfold_sites(text: str, active_abs_start: int, occupancies: dict[str, float]) -> list[Site]:
    sites: list[Site] = []
    for raw in text.splitlines():
        parts = raw.split()
        if len(parts) != 10 or parts[0] not in {"CREB1", "CREM", "ATF1", "ATF7"}:
            continue
        tf = parts[0]
        idx = int(parts[1])
        start_abs = int(parts[4])
        end_abs = int(parts[5])
        key = f"{tf}{idx}"
        sites.append(
            Site(
                tf=tf,
                start=start_abs - active_abs_start,
                end=end_abs - active_abs_start + 1,
                score=float(parts[6]),
                occupancy=occupancies.get(key, 0.0),
                strand=parts[9],
                index=idx,
            )
        )
    return sites


def run_unfold(unfold_exe: Path, xml_path: Path, section: str, args: list[str]) -> str:
    env = os.environ.copy()
    env["PATH"] = r"C:\msys64\ucrt64\bin;" + env.get("PATH", "")
    completed = subprocess.run(
        [str(unfold_exe), "--section", section, *args, "-i", str(xml_path)],
        check=True,
        capture_output=True,
        text=True,
        env=env,
    )
    return completed.stdout


def load_unfold_sites(
    unfold_exe: Path,
    xml_path: Path,
    section: str,
    genes: list[str],
    active_start: int,
) -> dict[str, list[Site]]:
    active_abs_start = LEFT_BOUND + active_start
    out: dict[str, list[Site]] = {}
    for gene in genes:
        occ_text = run_unfold(unfold_exe, xml_path, section, ["--occupancy", "--gene", gene])
        site_text = run_unfold(unfold_exe, xml_path, section, ["--sites", "--gene", gene])
        out[gene] = parse_unfold_sites(site_text, active_abs_start, parse_unfold_occupancy(occ_text))
    return out


def pwm_score(seq: str, matrix: dict[str, list[float]]) -> float:
    return sum(matrix[base][idx] for idx, base in enumerate(seq))


def scan_tfbs(seq: str, pwm: dict[str, dict[str, list[float]]], params: dict[str, TFParam]) -> list[Site]:
    sites: list[Site] = []
    for tf in TFS:
        matrix = pwm[tf]
        par = params[tf]
        motif_len = len(matrix["A"])
        maxscore = sum(max(matrix[base][idx] for base in BASES) for idx in range(motif_len))

        for start in range(0, len(seq) - motif_len + 1):
            window = seq[start : start + motif_len]
            candidates = [
                (pwm_score(window, matrix), "+"),
                (pwm_score(window.translate(RC)[::-1], matrix), "-"),
            ]
            for score, strand in candidates:
                if score < par.threshold:
                    continue
                k_exp = math.exp(-((maxscore - score) / par.lam))
                occupancy = (par.kmax * k_exp) / (1.0 + par.kmax * k_exp)
                sites.append(Site(tf, start, start + motif_len, score, occupancy, strand))
    return sites


def mutate_cre(seq: str, pos: int, base: str) -> str:
    return seq[:pos] + base + seq[pos + 1 :]


def mutate_cre_label(seq: str, pos_label: int, base: str) -> str:
    return mutate_cre(seq, label_to_index(pos_label), base)


def variant_lookup(variants, base: str) -> dict[int, dict]:
    return {v["pos"]: v for v in variants if v["base"] == base}


def pearson_r(x: list[float], y: list[float]) -> float:
    if len(x) < 2:
        return float("nan")
    a = np.asarray(x)
    b = np.asarray(y)
    return float(np.corrcoef(a, b)[0, 1])


def plot_activity_pair(ax_top, ax_bottom, variants, wt_expr, rates, base: str, paper_style: bool = False):
    subset = sorted([v for v in variants if v["base"] == base], key=lambda v: v["pos"])
    x = [v["pos"] for v in subset]
    measured = [math.log2(v["expr"] / wt_expr) for v in subset]

    top_colors = "#0000ff" if paper_style else ["blue" if v >= 0 else "red" for v in measured]
    ax_top.bar(x, measured, width=0.72, color=top_colors, linewidth=0)
    style_activity_axis(ax_top, base, paper_style)
    for name, (start, end, _) in CRE_SITES.items():
        ax_top.text((start + end) / 2, 2.08, name, ha="center", va="bottom", fontsize=7)

    predicted = []
    pred_x = []
    wt_pred = rates.get("synCRE_Promega_0")
    if wt_pred:
        for v in subset:
            if v["name"] in rates:
                pred_x.append(v["pos"])
                predicted.append(math.log2(rates[v["name"]] / wt_pred))
        bottom_colors = "#ff0000" if paper_style else ["blue" if v >= 0 else "red" for v in predicted]
        ax_bottom.bar(pred_x, predicted, width=0.72, color=bottom_colors, linewidth=0)
        common = {v["pos"]: math.log2(rates[v["name"]] / wt_pred) for v in subset if v["name"] in rates}
        meas_common = [math.log2(v["expr"] / wt_expr) for v in subset if v["pos"] in common]
        pred_common = [common[v["pos"]] for v in subset if v["pos"] in common]
        if pred_common:
            r = pearson_r(meas_common, pred_common)
            rms = float(np.sqrt(np.mean((np.asarray(meas_common) - np.asarray(pred_common)) ** 2)))
            metric_text = (
                rf"R = {r:.2f} R$^2$ = {r*r:.2f} rms = {rms:.2f}"
                if paper_style else f"R = {r:.2f} R2 = {r*r:.2f} rms = {rms:.2f}"
            )
            ax_bottom.text(0.98, 0.92, metric_text, transform=ax_bottom.transAxes,
                           ha="right", va="top", fontsize=8)
    else:
        ax_bottom.text(0.5, 0.5, "model .rates not found", ha="center", va="center", fontsize=8)

    highlight = CRE1_VARIANTS_A + CRE4_VARIANTS_A if base == "A" else CRE1_VARIANTS_T + CRE4_VARIANTS_T
    for pos, label_base in highlight:
        ax_bottom.text(pos, -1.72 if paper_style else -2.12, f"{pos}{label_base}",
                       rotation=0 if paper_style else 90, ha="center", va="top" if paper_style else "bottom",
                       fontsize=4.6 if paper_style else 5.5)

    style_activity_axis(ax_bottom, base, paper_style)
    ax_bottom.set_xlabel("position")
    if paper_style:
        ax_top.tick_params(labelbottom=False)


def style_activity_axis(ax, base: str, paper_style: bool = False):
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_xlim(-3, 87)
    ax.set_ylim(-2.2, 2.2)
    ax.set_ylabel("Δactivity (log2)" if paper_style else "Dactivity (log2)", fontsize=8)
    ax.grid(True, axis="both" if paper_style else "y", color="#dddddd", linewidth=0.55)
    for name, (start, end, color) in CRE_SITES.items():
        ax.axvspan(start, end, color=color, alpha=0.28, zorder=0)
    ax.text(0.04, 0.83, f"→{base}" if paper_style else f"->{base}", transform=ax.transAxes, fontsize=11)
    if paper_style:
        ax.set_xticks(np.arange(0, 81, 10))
        ax.tick_params(labelsize=7, length=3, width=0.7)
        for spine in ax.spines.values():
            spine.set_linewidth(0.7)


def plot_tfbs_track(ax, sites: list[Site], title: str | None = None, paper_style: bool = False,
                    show_fraction_legend: bool = False):
    ymap = {"ATF7": 3, "ATF1": 2, "CREM": 1, "CREB1": 0}
    ax.set_xlim(-1, 88)
    ax.set_ylim(-0.8, 4.2)
    ax.set_yticks([ymap[tf] for tf in TFS])
    ax.set_yticklabels(TFS, fontsize=8)
    ax.set_xticks([])
    for site in sites:
        alpha = min(0.82, max(0.02 if paper_style else 0.12, site.occupancy))
        ax.add_patch(
            patches.Rectangle(
                (site.start, ymap[site.tf] - 0.35),
                site.end - site.start,
                0.7,
                facecolor=TF_COLORS[site.tf],
                edgecolor=TF_COLORS[site.tf],
                alpha=alpha,
                linewidth=0.9 if paper_style else 1.0,
            )
        )
        if paper_style:
            ax.add_patch(patches.Rectangle(
                (site.start, ymap[site.tf] - 0.35), site.end - site.start, 0.7,
                facecolor="none", edgecolor=TF_COLORS[site.tf], linewidth=0.9,
            ))
    for name, (start, end, color) in CRE_SITES.items():
        ax.add_patch(patches.Rectangle((start, -0.72), end - start, 0.26, color=color, alpha=0.75))
    if paper_style:
        ax.axhline(-0.44, color="black", linewidth=0.65)
        ax.add_patch(patches.Rectangle((0, -0.72), 3, 0.26, color="#bfbfbf", alpha=0.7))
        ax.add_patch(patches.Rectangle((83, -0.72), 4, 0.26, color="#bfbfbf", alpha=0.7))
        ax.tick_params(axis="y", labelsize=7, length=3)
        for spine in ax.spines.values():
            spine.set_linewidth(0.7)
        if show_fraction_legend:
            x0, x1, y0, y1 = 1.018, 1.085, 0.12, 0.88
            bands = 36
            for band in range(bands):
                band_y0 = y0 + (y1 - y0) * band / bands
                band_y1 = y0 + (y1 - y0) * (band + 1) / bands
                band_x0 = x0 + (x1 - x0) * (band_y0 - y0) / (y1 - y0)
                band_x1 = x0 + (x1 - x0) * (band_y1 - y0) / (y1 - y0)
                shade = 1.0 - 0.82 * (band + 0.5) / bands
                ax.add_patch(patches.Polygon(
                    [[x0, band_y0], [band_x0, band_y0], [band_x1, band_y1], [x0, band_y1]],
                    closed=True, transform=ax.transAxes, facecolor=(shade, shade, shade),
                    edgecolor="none", clip_on=False,
                ))
            ax.add_patch(patches.Polygon(
                [[x0, y0], [x1, y1], [x0, y1]], closed=True, transform=ax.transAxes,
                facecolor="none", edgecolor="black", linewidth=0.5, clip_on=False,
            ))
            ax.text(1.09, 0.89, "High F", transform=ax.transAxes, fontsize=6.5,
                    ha="left", va="center", clip_on=False)
            ax.text(1.09, 0.10, "Low F", transform=ax.transAxes, fontsize=6.5,
                    ha="left", va="center", clip_on=False)
    if title:
        ax.set_title(title, fontsize=9)


def plot_variant_box(ax, wt_cre: str, pwm, params, pos: int, base: str, region: tuple[int, int],
                     sites_override=None, paper_style: bool = False):
    pos_index = label_to_index(pos)
    seq = mutate_cre_label(wt_cre, pos, base)
    source_sites = sites_override if sites_override is not None else scan_tfbs(seq, pwm, params)
    sites = [s for s in source_sites if s.end >= region[0] and s.start <= region[1]]
    ymap = {"ATF7": 3, "ATF1": 2, "CREM": 1, "CREB1": 0}
    left, right = region
    ax.set_xlim(left, right)
    ax.set_ylim(-1.00 if paper_style else -1.2, 4.2)
    ax.set_yticks([ymap[tf] for tf in TFS])
    ax.set_yticklabels(TFS, fontsize=7)
    ax.set_xticks([])
    for site in sites:
        alpha = min(0.82, max(0.02 if paper_style else 0.12, site.occupancy))
        ax.add_patch(
            patches.Rectangle(
                (site.start, ymap[site.tf] - 0.32),
                site.end - site.start,
                0.64,
                facecolor=TF_COLORS[site.tf],
                edgecolor=TF_COLORS[site.tf],
                alpha=alpha,
            )
        )
        if paper_style:
            ax.add_patch(patches.Rectangle(
                (site.start, ymap[site.tf] - 0.32), site.end - site.start, 0.64,
                facecolor="none", edgecolor=TF_COLORS[site.tf], linewidth=0.9,
            ))
    if paper_style:
        feature_y = -0.86
        feature_h = 0.24
        seq_y = -1.23
        outline_y = -1.43
        outline_h = 0.36
        for _, (feature_start, feature_end, color) in CRE_SITES.items():
            clipped_start = max(left, feature_start)
            clipped_end = min(right, feature_end)
            if clipped_start < clipped_end:
                bar_start = clipped_start
                bar_end = clipped_end
                if color == "#ff6b6b":
                    inset = min(0.9, (clipped_end - clipped_start) * 0.16)
                    bar_start += inset
                    bar_end -= inset
                ax.add_patch(patches.Rectangle(
                    (bar_start, feature_y), bar_end - bar_start, feature_h,
                    color=color, alpha=0.85, zorder=3,
                ))
        ax.axhline(-0.58, color="black", linewidth=0.65)
        for idx in range(left, right):
            ax.text(idx + 0.5, seq_y, seq[idx], ha="center", va="center", fontsize=7,
                    family="monospace", bbox=dict(facecolor="#ffd42a" if idx == pos_index else "none",
                                                   edgecolor="none", pad=0.05),
                    zorder=6, clip_on=False)
        for _, (feature_start, feature_end, color) in CRE_SITES.items():
            if color == "#ff6b6b" and feature_start < right and feature_end > left:
                box_start = max(left, feature_start)
                box_end = min(right, feature_end)
                inset = min(0.9, (box_end - box_start) * 0.16)
                bar_start = box_start + inset
                bar_end = box_end - inset
                ax.plot(
                    [bar_start, box_start],
                    [feature_y, outline_y + outline_h],
                    color="black",
                    linewidth=0.55,
                    zorder=4,
                    clip_on=False,
                )
                ax.plot(
                    [bar_end, box_end],
                    [feature_y, outline_y + outline_h],
                    color="black",
                    linewidth=0.55,
                    zorder=4,
                    clip_on=False,
                )
                ax.add_patch(patches.Rectangle(
                    (box_start, outline_y), box_end - box_start, outline_h,
                    facecolor="none", edgecolor="#e31a1c", linewidth=0.8,
                    zorder=5, clip_on=False,
                ))
        ax.tick_params(axis="y", labelsize=7, length=3)
        for spine in ax.spines.values():
            spine.set_linewidth(0.7)
    else:
        ax.add_patch(patches.Rectangle((pos_index, -0.95), 1, 0.28, color="#ffd42a"))
        label_seq = seq[left:right]
        ax.text((left + right) / 2, -0.48, label_seq, ha="center", va="center", fontsize=7, family="monospace")
    ax.text(
        0.965,
        0.93,
        f"{pos}{base}",
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=9,
        zorder=10,
    )


def dda_by_site(seq: str, pwm, params, region: tuple[int, int], sites_override=None) -> list[tuple[str, float]]:
    source_sites = sites_override if sites_override is not None else scan_tfbs(seq, pwm, params)
    sites = [s for s in source_sites if s.end >= region[0] and s.start <= region[1]]
    return [(s.tf, params[s.tf].coef * s.occupancy) for s in sites]


def plot_dda(ax, wt_cre, pwm, params, groups, title, site_lookup=None, paper_style: bool = False,
             substitution_base: str | None = None):
    labels: list[str] = []
    seqs: list[str] = []
    genes: list[str] = []
    regions: list[tuple[int, int]] = []
    positions: list[float] = []
    group_centers: list[tuple[str, float]] = []
    group_ranges: list[tuple[float, float]] = []
    cursor = 0.0

    for group_name, variants_def, region in groups:
        start_cursor = cursor
        group_labels = ["WT"] + [f"{pos}{base}" for pos, base in variants_def]
        group_seqs = [wt_cre] + [mutate_cre_label(wt_cre, pos, base) for pos, base in variants_def]
        group_genes = ["synCRE_Promega_0"] + [variant_gene(pos, base) for pos, base in variants_def]
        for label, seq, gene in zip(group_labels, group_seqs, group_genes):
            labels.append(label)
            seqs.append(seq)
            genes.append(gene)
            regions.append(region)
            positions.append(cursor)
            cursor += 1.0
        group_centers.append((group_name, (start_cursor + cursor - 1.0) / 2.0))
        group_ranges.append((start_cursor, cursor - 1.0))
        cursor += 0.8

    bottoms = np.zeros(len(seqs))
    for tf in ["CREB1", "CREM", "ATF1", "ATF7"]:
        vals = []
        for seq, region, gene in zip(seqs, regions, genes):
            override = site_lookup.get(gene) if site_lookup else None
            vals.append(sum(value for site_tf, value in dda_by_site(seq, pwm, params, region, override) if site_tf == tf))
        ax.bar(positions, vals, bottom=bottoms, color=TF_COLORS[tf], edgecolor="black", linewidth=0.4, label=tf)
        bottoms += np.asarray(vals)
    ax.set_xticks(positions)
    ax.set_xticklabels(labels, fontsize=7)
    ax.set_ylabel("ΔΔA" if paper_style else "DDA", fontsize=8)
    ax.set_title(title, fontsize=10 if paper_style else 8, loc="left" if paper_style else "center",
                 fontweight="bold" if paper_style else "normal")
    top = max(bottoms) if len(bottoms) else 1.0
    for (group_name, center), (group_start, group_end) in zip(group_centers, group_ranges):
        ax.add_patch(patches.Rectangle((group_start - 0.42, top * 1.04), group_end - group_start + 0.84,
                                       top * 0.05, color="#cccccc", clip_on=False))
        ax.text(center, top * 1.12, group_name, ha="center", va="bottom", fontsize=7)
        if paper_style:
            indices = [idx for idx, position in enumerate(positions) if group_start <= position <= group_end]
            ax.plot([positions[idx] for idx in indices], [bottoms[idx] for idx in indices],
                    color="#333333", linewidth=0.65, zorder=5)
    if paper_style and substitution_base:
        ax.text(0.01, 0.93, f"→{substitution_base}", transform=ax.transAxes,
                ha="left", va="top", fontsize=9)
        ax.tick_params(labelsize=7, length=3, width=0.7)
        for spine in ax.spines.values():
            spine.set_linewidth(0.7)
    ax.set_ylim(0, top * 1.25 if top > 0 else 1)


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def export_panel_data(data_dir, variants, wt_expr, rates, params, site_lookup, wt_cre, pwm):
    data_dir.mkdir(parents=True, exist_ok=True)
    wt_rate = rates.get("synCRE_Promega_0")
    for base in ("A", "T"):
        rows = []
        for variant in sorted((v for v in variants if v["base"] == base), key=lambda v: v["pos"]):
            predicted = rates.get(variant["name"])
            rows.append({
                "position": variant["pos"],
                "base": base,
                "gene": variant["name"],
                "measured_expression": variant["expr"],
                "measured_log2_change": math.log2(variant["expr"] / wt_expr),
                "predicted_rate": "" if predicted is None else predicted,
                "predicted_log2_change": "" if predicted is None or not wt_rate else math.log2(predicted / wt_rate),
            })
        write_tsv(
            data_dir / f"activity_{base}.tsv",
            ["position", "base", "gene", "measured_expression", "measured_log2_change", "predicted_rate", "predicted_log2_change"],
            rows,
        )

    site_rows = []
    for gene, sites in site_lookup.items():
        for site in sites:
            site_rows.append({
                "gene": gene, "TF": site.tf, "site_index": site.index,
                "start": site.start, "end": site.end, "strand": site.strand,
                "PWM_score": site.score, "fractional_occupancy": site.occupancy,
            })
    write_tsv(
        data_dir / "occupancy_sites_selected.tsv",
        ["gene", "TF", "site_index", "start", "end", "strand", "PWM_score", "fractional_occupancy"],
        site_rows,
    )

    definitions = {
        "A": [("CRE1", CRE1_VARIANTS_A, CRE_SITES["CRE1"][:2]), ("CRE4", CRE4_VARIANTS_A, CRE_SITES["CRE4"][:2])],
        "T": [("CRE1", CRE1_VARIANTS_T, CRE_SITES["CRE1"][:2]), ("CRE4", CRE4_VARIANTS_T, CRE_SITES["CRE4"][:2])],
    }
    for base, groups in definitions.items():
        dda_rows = []
        for group_name, variants_def, region in groups:
            items = [("WT", "synCRE_Promega_0", wt_cre)] + [
                (f"{pos}{variant_base}", variant_gene(pos, variant_base), mutate_cre_label(wt_cre, pos, variant_base))
                for pos, variant_base in variants_def
            ]
            for label, gene, seq in items:
                override = site_lookup.get(gene)
                values = dda_by_site(seq, pwm, params, region, override)
                for tf in ["CREB1", "CREM", "ATF1", "ATF7"]:
                    dda_rows.append({
                        "group": group_name, "variant": label, "gene": gene,
                        "TF": tf, "DDA": sum(value for site_tf, value in values if site_tf == tf),
                    })
        write_tsv(data_dir / f"DDA_{base}.tsv", ["group", "variant", "gene", "TF", "DDA"], dda_rows)

    comparable = [(wt_expr, rates.get("synCRE_Promega_0"))]
    comparable.extend((variant["expr"], rates.get(variant["name"])) for variant in variants)
    comparable = [(observed, predicted) for observed, predicted in comparable if predicted is not None]
    metrics = {
        "forward_sse_from_rounded_model5_parameters": sum((observed - predicted) ** 2 for observed, predicted in comparable),
        "comparable_genes": len(comparable),
        "wt_measured_expression": wt_expr,
        "wt_predicted_rate": rates.get("synCRE_Promega_0"),
    }
    for base in ("A", "T"):
        pairs = [
            (math.log2(variant["expr"] / wt_expr), math.log2(rates[variant["name"]] / rates["synCRE_Promega_0"]))
            for variant in variants
            if variant["base"] == base and variant["name"] in rates and rates.get("synCRE_Promega_0")
        ]
        metrics[f"pearson_r_{base}"] = pearson_r([pair[0] for pair in pairs], [pair[1] for pair in pairs])
        metrics[f"variants_{base}"] = len(pairs)
    (data_dir / "summary_metrics.json").write_text(json.dumps(metrics, indent=2), encoding="utf-8")


def main():
    package_dir = Path(__file__).resolve().parent
    project_dir = package_dir.parent

    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=package_dir / "source_data")
    parser.add_argument("--rates", type=Path, default=package_dir / "model" / "best_fit.output.rates")
    parser.add_argument("--xml", type=Path, default=package_dir / "model" / "best_fit.xml")
    parser.add_argument("--unfold", type=Path, default=project_dir / "portable_4tf_scrambled_20260611" / "bin" / "unfold.exe")
    parser.add_argument("--out", type=Path, default=package_dir / "results" / "Figure5_paper_detail_refined.png")
    parser.add_argument("--pdf-out", type=Path, default=package_dir / "results" / "Figure5_paper_detail_refined.pdf")
    parser.add_argument("--data-dir", type=Path, default=package_dir / "results" / "panel_data_refined")
    parser.add_argument("--paper-style", dest="paper_style", action="store_true", default=True,
                        help="Use a layout and visual encoding closer to the published Figure 5")
    parser.add_argument("--basic-style", dest="paper_style", action="store_false",
                        help="Use the earlier broad diagnostic layout instead of the paper-style layout")
    args = parser.parse_args()

    mmc2 = find_one("*mmc2.xlsx", args.root)
    mmc3 = find_one("*mmc3.xlsx", args.root)
    mmc4 = find_one("*mmc4.xlsx", args.root)
    rates_path = args.rates or next(iter(glob.glob(str(args.root / "**" / "*.rates"), recursive=True)), None)

    params = load_mmc2_params(mmc2)
    if args.xml.exists():
        try:
            params = load_xml_output_params(args.xml)
        except Exception as exc:
            print(f"Warning: using mmc2 params because XML params could not be loaded: {exc}")
    pwm = load_pwm(mmc3)
    _, wt_cre, wt_expr, active_start, variants = load_single_hit(mmc4)
    rates = load_rates(Path(rates_path)) if rates_path else {}
    selected_defs = CRE1_VARIANTS_A + CRE4_VARIANTS_A + CRE1_VARIANTS_T + CRE4_VARIANTS_T
    selected_genes = ["synCRE_Promega_0"] + [variant_gene(pos, base) for pos, base in selected_defs]
    site_lookup = {}
    if args.unfold.exists() and args.xml.exists():
        xml_root = ET.parse(args.xml).getroot()
        unfold_section = "Output" if xml_root.find("Output") is not None else "Input"
        print(f"Using XML section {unfold_section} for TFBS and occupancy")
        site_lookup = load_unfold_sites(args.unfold, args.xml, unfold_section, selected_genes, active_start)
    wt_sites = site_lookup.get("synCRE_Promega_0", scan_tfbs(wt_cre, pwm, params))
    if args.data_dir:
        export_panel_data(args.data_dir, variants, wt_expr, rates, params, site_lookup, wt_cre, pwm)

    if args.paper_style:
        fig = plt.figure(figsize=(10.6, 8.9), dpi=200)
        gs = fig.add_gridspec(4, 2, height_ratios=[2.05, 0.72, 3.0, 1.05],
                              hspace=0.31, wspace=0.28)
        activity_left = gs[0, 0].subgridspec(2, 1, hspace=0.08)
        activity_right = gs[0, 1].subgridspec(2, 1, hspace=0.08)
        ax_a1 = fig.add_subplot(activity_left[0, 0])
        ax_a2 = fig.add_subplot(activity_left[1, 0], sharex=ax_a1)
        ax_g1 = fig.add_subplot(activity_right[0, 0])
        ax_g2 = fig.add_subplot(activity_right[1, 0], sharex=ax_g1)
    else:
        fig = plt.figure(figsize=(11, 9), dpi=160)
        gs = fig.add_gridspec(6, 6, height_ratios=[1.0, 1.0, 0.7, 1.05, 1.05, 0.8],
                              hspace=0.42, wspace=0.45)
        ax_a1 = fig.add_subplot(gs[0, 0:3])
        ax_a2 = fig.add_subplot(gs[1, 0:3], sharex=ax_a1)
        ax_g1 = fig.add_subplot(gs[0, 3:6])
        ax_g2 = fig.add_subplot(gs[1, 3:6], sharex=ax_g1)

    plot_activity_pair(ax_a1, ax_a2, variants, wt_expr, rates, "A", args.paper_style)
    plot_activity_pair(ax_g1, ax_g2, variants, wt_expr, rates, "T", args.paper_style)
    ax_a1.set_title("A", loc="left", fontweight="bold", fontsize=11 if args.paper_style else None)
    ax_g1.set_title("G", loc="left", fontweight="bold", fontsize=11 if args.paper_style else None)

    if args.paper_style:
        ax_b = fig.add_subplot(gs[1, 0])
        ax_h = fig.add_subplot(gs[1, 1])
    else:
        ax_b = fig.add_subplot(gs[2, 0:3])
        ax_h = fig.add_subplot(gs[2, 3:6])
    plot_tfbs_track(ax_b, wt_sites, paper_style=args.paper_style)
    plot_tfbs_track(ax_h, wt_sites, paper_style=args.paper_style, show_fraction_legend=args.paper_style)
    ax_b.set_title("B", loc="left", fontweight="bold", fontsize=11 if args.paper_style else None)
    ax_h.set_title("H", loc="left", fontweight="bold", fontsize=11 if args.paper_style else None)

    variant_definitions = [
        ("C", CRE1_VARIANTS_A, (7, 22)),
        ("D", CRE4_VARIANTS_A, (63, 81)),
        ("I", CRE1_VARIANTS_T, (7, 22)),
        ("J", CRE4_VARIANTS_T, (63, 81)),
    ]
    if args.paper_style:
        variant_left = gs[2, 0].subgridspec(3, 2, hspace=0.35, wspace=0.32)
        variant_right = gs[2, 1].subgridspec(3, 2, hspace=0.35, wspace=0.32)
        for definition_index, (panel, variants_def, region) in enumerate(variant_definitions):
            local_col = definition_index % 2
            local_grid = variant_left if definition_index < 2 else variant_right
            for idx, (pos, base) in enumerate(variants_def):
                ax = fig.add_subplot(local_grid[idx, local_col])
                gene = variant_gene(pos, base)
                plot_variant_box(ax, wt_cre, pwm, params, pos, base, region,
                                 site_lookup.get(gene), paper_style=True)
                if idx == 0:
                    ax.set_title(panel, loc="left", fontweight="bold", fontsize=11)
    else:
        for col, (panel, variants_def, region) in enumerate(variant_definitions):
            base_col = col if col < 2 else col + 1
            outer = gs[3:5, base_col].subgridspec(3, 1, hspace=0.15)
            for idx, (pos, base) in enumerate(variants_def):
                ax = fig.add_subplot(outer[idx, 0])
                gene = variant_gene(pos, base)
                plot_variant_box(ax, wt_cre, pwm, params, pos, base, region, site_lookup.get(gene))
                if idx == 0:
                    ax.set_title(panel, loc="left", fontweight="bold")

    if args.paper_style:
        bottom_left = gs[3, 0].subgridspec(1, 2, width_ratios=[0.72, 2.28], wspace=0.48)
        ax_e = fig.add_subplot(bottom_left[0, 0])
        ax_f = fig.add_subplot(bottom_left[0, 1])
        ax_k = fig.add_subplot(gs[3, 1])
    else:
        ax_e = fig.add_subplot(gs[5, 0])
        ax_f = fig.add_subplot(gs[5, 1:3])
        ax_k = fig.add_subplot(gs[5, 3:6])

    coef_tfs = ["CREB1", "CREM", "ATF1", "ATF7"]
    ax_e.bar(range(len(coef_tfs)), [math.log10(params[tf].coef) for tf in coef_tfs],
             color=[TF_COLORS[tf] for tf in coef_tfs], edgecolor="none")
    ax_e.set_xticks(range(len(coef_tfs)))
    ax_e.set_xticklabels(coef_tfs, rotation=45, ha="right", fontsize=7)
    ax_e.set_ylabel("log10(coef)", fontsize=8)
    ax_e.set_title("E", loc="left", fontweight="bold", fontsize=11 if args.paper_style else None)
    if args.paper_style:
        ax_e.axhline(0, color="black", linewidth=0.65)
        ax_e.tick_params(labelsize=7, length=3, width=0.7)
        for spine in ax_e.spines.values():
            spine.set_linewidth(0.7)

    plot_dda(
        ax_f,
        wt_cre,
        pwm,
        params,
        [
            ("CRE1", CRE1_VARIANTS_A, CRE_SITES["CRE1"][:2]),
            ("CRE4", CRE4_VARIANTS_A, CRE_SITES["CRE4"][:2]),
        ],
        "F",
        site_lookup,
        args.paper_style,
        "A" if args.paper_style else None,
    )
    plot_dda(
        ax_k,
        wt_cre,
        pwm,
        params,
        [
            ("CRE1", CRE1_VARIANTS_T, CRE_SITES["CRE1"][:2]),
            ("CRE4", CRE4_VARIANTS_T, CRE_SITES["CRE4"][:2]),
        ],
        "K",
        site_lookup,
        args.paper_style,
        "T" if args.paper_style else None,
    )
    ax_k.legend(fontsize=6, ncols=4, frameon=False, loc="upper right",
                bbox_to_anchor=(1.0, 1.17 if args.paper_style else 1.0))

    if args.paper_style:
        fig.subplots_adjust(left=0.07, right=0.93, top=0.975, bottom=0.07)

    fig.savefig(args.out, bbox_inches="tight")
    print(f"Saved {args.out}")
    if args.pdf_out:
        fig.savefig(args.pdf_out, bbox_inches="tight")
        print(f"Saved {args.pdf_out}")


if __name__ == "__main__":
    main()
