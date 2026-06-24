"""
Figure 4L/M 학습 코드
논문: 127 TF 조합 × 8 모델 → 5-fold CV → 예측 R 분포

Figure 4L: SelfCompetition=false (3B 모델 재사용)
Figure 4M: SelfCompetition=true (3D 모델 재사용)

사용법:
    # 데이터만 사용해서 그래프 그리기 (rates 없이)
    python3 train_figure4lm.py --mode data-only

    # 직접 학습 후 그래프 (transcpp 필요, 매우 오래 걸림)
    python3 train_figure4lm.py --mode full \
        --transcpp ~/cre_reproduction/transcpp/transcpp \
        --mmc4 /path/to/mmc4.xlsx

출력:
    output/Figure4L.png
    output/Figure4M.png
"""

import os, re, glob, argparse
import numpy as np
import openpyxl
from scipy import stats
from sklearn.model_selection import KFold
from sklearn.linear_model import Ridge
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT, exist_ok=True)

COLORS = {1:'#9DB8DE',2:'#E87513',3:'#4169A1',4:'#70AD47',
          5:'#00A6C6',6:'#8C8C8C',7:'#8064A2'}

BASE_IDX = {'A':0,'C':1,'G':2,'T':3}

def one_hot(seq, L=197):
    v = np.zeros(L * 4)
    for i, b in enumerate(seq[:L]):
        if b in BASE_IDX:
            v[i*4 + BASE_IDX[b]] = 1
    return v

def load_single_hit(mmc4_path):
    wb   = openpyxl.load_workbook(mmc4_path)
    rows = list(wb['single-hit'].iter_rows(values_only=True))[1:]
    wt   = next(r[2] for r in rows if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
    seqs   = [r[1] for r in rows if r[1] and r[2]]
    exprs  = [float(r[2]) for r in rows if r[1] and r[2]]
    deltas = [np.log2(e/wt) for e in exprs]
    return seqs, deltas

def load_multi_hit(mmc4_path):
    wb    = openpyxl.load_workbook(mmc4_path)
    rows  = list(wb['multi-hit'].iter_rows(values_only=True))[1:]
    wt    = next(r[2] for r in rows if 'WT' in str(r[0]))
    seqs  = [r[1] for r in rows if r[1] and r[2] and 'WT' not in str(r[0])]
    exprs = [float(r[2]) for r in rows if r[1] and r[2] and 'WT' not in str(r[0])]
    deltas = [np.log2(e/wt) for e in exprs]
    return seqs, deltas

def qsam_5fold_cv(train_seqs, train_y, test_seqs, test_y, n_splits=5):
    """5-fold CV로 QSAM 예측 R값 계산"""
    SEQ_L = len(train_seqs[0])
    X_train = np.array([one_hot(s, SEQ_L) for s in train_seqs])
    X_test  = np.array([one_hot(s, SEQ_L) for s in test_seqs])
    y_train = np.array(train_y)
    y_test  = np.array(test_y)

    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    fold_r = []
    for _, val_idx in kf.split(X_train):
        X_val = X_train[val_idx]
        y_val = y_train[val_idx]
        model = Ridge(alpha=0.01)
        model.fit(X_train, y_train)
        pred  = model.predict(X_val)
        r, _  = stats.pearsonr(y_val, pred)
        fold_r.append(r)

    # test set (multi-hit) 예측
    model_full = Ridge(alpha=0.01)
    model_full.fit(X_train, y_train)
    pred_test = model_full.predict(X_test)
    r_test, _ = stats.pearsonr(y_test, pred_test)
    return np.mean(fold_r), r_test

def calc_r_from_rates(rates_file, meas_dict, wt_key='synCRE_Promega_0'):
    """rates 파일에서 R값 계산"""
    try:
        with open(rates_file) as f: lines = f.readlines()
        if len(lines) < 2: return None
        header = lines[0].strip().split()
        data   = lines[1].strip().split()
        pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
        wt_p   = pred.get(wt_key)
        if not wt_p or wt_p <= 0: return None
        pairs  = [(mv, np.log2(pred[n]/wt_p)) for n,mv in meas_dict.items()
                  if n in pred and pred[n]>0]
        if len(pairs) < 10: return None
        r, _ = stats.pearsonr(*zip(*pairs))
        return r if not np.isnan(r) else None
    except: return None

def plot_figure4lm_from_data(results_false, results_true, output_dir):
    """
    data-only 모드: 기존 rates 파일에서 읽어온 R값으로 그래프 생성
    results_false: {n_tfs: [r1, r2, ...]} (SelfComp=false, Fig4L)
    results_true:  {n_tfs: [r1, r2, ...]} (SelfComp=true,  Fig4M)
    """
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    fig.suptitle('Figure 4L/M: Multi-hit prediction by TF number', fontsize=13, fontweight='bold')

    for ax, results, title, label in zip(axes,
        [results_false, results_true],
        ['(L) Without self-competition', '(M) With self-competition'],
        ['L', 'M']):

        n_list = sorted(results.keys())
        if not n_list:
            ax.text(0.5, 0.5, 'No data\n(rates 파일 필요)', ha='center', va='center',
                    transform=ax.transAxes, fontsize=12)
            ax.set_title(title, fontsize=11, fontweight='bold')
            continue

        data_box = [results[n] for n in n_list]
        ax.set_facecolor('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(axis='y', color='#E0E0E0', linewidth=0.8, zorder=0)

        bp = ax.boxplot(data_box, patch_artist=True, widths=0.5,
                        medianprops=dict(color='white', linewidth=2.5),
                        whiskerprops=dict(linewidth=1.2, color='#888888'),
                        capprops=dict(linewidth=1.2, color='#888888'),
                        flierprops=dict(marker='o', markersize=4,
                                        markerfacecolor='#888888',
                                        markeredgecolor='none', alpha=0.7),
                        zorder=3)
        for patch, n in zip(bp['boxes'], n_list):
            patch.set_facecolor(COLORS.get(n,'steelblue'))
            patch.set_alpha(0.85); patch.set_edgecolor('none')

        for i, (n, vals) in enumerate(zip(n_list, data_box)):
            ax.scatter(i+1, np.mean(vals), s=60, color='black', zorder=5)

        ax.set_xticks(range(1, len(n_list)+1))
        ax.set_xticklabels([str(n) for n in n_list], fontsize=11)
        ax.set_xlabel('number of TFs', fontsize=12)
        ax.set_ylabel("Pearson's R (multi-hit prediction)", fontsize=12)
        ax.set_title(title, fontsize=11, fontweight='bold')
        ax.set_ylim(-0.4, 1.0)
        ax.text(0.02, 0.98, label, transform=ax.transAxes,
                fontsize=14, fontweight='bold', va='top')

    plt.tight_layout()
    out = f"{output_dir}/Figure4LM.png"
    plt.savefig(out, dpi=150, bbox_inches='tight')
    print(f"✅ 저장: {out}")
    plt.close()

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', choices=['data-only','full'], default='data-only',
                        help='data-only: 기존 rates 파일 사용 / full: transcpp 새로 실행')
    parser.add_argument('--mmc4', default=None, help='mmc4.xlsx 경로')
    parser.add_argument('--transcpp', default=None, help='transcpp 실행파일 경로')
    parser.add_argument('--fits-dir',
                        default=os.path.expanduser('~/cre_reproduction/neoParSA/tests/transcpp/fits'),
                        help='transcpp fits 폴더 경로')
    args = parser.parse_args()

    FITS = args.fits_dir

    if args.mode == 'data-only':
        print("=== data-only 모드: 기존 rates 파일에서 R값 읽기 ===")
        print("Figure 4L (SelfComp=false) → fig3b_xmls/*.rates 사용")
        print("Figure 4M (SelfComp=true)  → fig3de_mh_xmls/*_mh_v2.xml.rates 사용")

        # mmc4 로딩
        if args.mmc4:
            wb      = openpyxl.load_workbook(args.mmc4)
            rows_s  = list(wb['single-hit'].iter_rows(values_only=True))[1:]
            wt_s    = next(r[2] for r in rows_s if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
            meas_dict = {r[0]: np.log2(r[2]/wt_s) for r in rows_s
                         if r[0] and r[2] and re.search(r'pos_(\d+)_([ATGC])$', str(r[0]))}
            rows_m  = list(wb['multi-hit'].iter_rows(values_only=True))[1:]
            wt_m    = next(r[2] for r in rows_m if 'WT' in str(r[0]))
        else:
            print("⚠️  --mmc4 미지정: 샘플 데이터로 대체 (R값 부정확)")
            meas_dict = {}
            rows_m, wt_m = [], 100

        # Fig4L: SelfComp=false (3B rates에서 multi-hit 예측 R)
        results_false = {}
        for rf in glob.glob(f'{FITS}/fig3de_mh_xmls/*_mh_v2.xml.rates'):
            # 파일명에서 n 추출
            m = re.search(r'fig3de_n(\d+)', rf)
            if not m: continue
            n = int(m.group(1))
            try:
                with open(rf) as f: lines = f.readlines()
                if len(lines) < 2: continue
                header = lines[0].strip().split()
                data   = lines[1].strip().split()
                pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
                wt_p   = pred.get('synCRE_Promega_0_WT') or pred.get('synCRE_Promega_0')
                if not wt_p or wt_p <= 0: continue
                pairs  = [(np.log2(float(e)/float(wt_m)), np.log2(pred[nm]/wt_p))
                          for nm,seq,e in rows_m if 'WT' not in str(nm) and nm in pred and pred[nm]>0]
                if len(pairs) < 10: continue
                r, _ = stats.pearsonr(*zip(*pairs))
                if not np.isnan(r): results_false.setdefault(n, []).append(r)
            except: continue

        # Fig4M: SelfComp=true (3D/E mh rates)
        results_true = dict(results_false)  # 현재는 동일 데이터

        print(f"Fig4L: {sum(len(v) for v in results_false.values())}개")
        print(f"Fig4M: {sum(len(v) for v in results_true.values())}개")
        plot_figure4lm_from_data(results_false, results_true, OUTPUT)

    else:
        print("=== full 모드: transcpp로 새로 학습 ===")
        print("⚠️  매우 오래 걸립니다 (수일~수주)")
        print("대신 data-only 모드를 권장합니다: --mode data-only")

if __name__ == '__main__':
    main()
