"""
Figure 2C 재현 스크립트
Kang & Kim 2024 (iScience)

사용법:
    cd ~/cre_reproduction/neoParSA/tests/transcpp/fits
    python3 /path/to/generate_figure2c.py

출력:
    Figure2C/output/Figure2C.png
"""

import os, re, subprocess
import numpy as np
import openpyxl
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# ===== 경로 설정 (사용자 환경에 맞게 수정) =====
BASE     = os.path.expanduser("~/Desktop/SB_final/ScienceDirect_files_26May2026_04-41-20.049")
FITS     = os.path.expanduser("~/cre_reproduction/neoParSA/tests/transcpp/fits")
TRANSCPP = os.path.expanduser("~/cre_reproduction/transcpp/transcpp")
MMC2     = f"{BASE}/1-s2.0-S2589004223028249-mmc2.xlsx"
MMC3     = f"{BASE}/1-s2.0-S2589004223028249-mmc3.xlsx"
MMC4     = f"{BASE}/1-s2.0-S2589004223028249-mmc4.xlsx"
OUTPUT   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT, exist_ok=True)

def load_single_hit():
    wb   = openpyxl.load_workbook(MMC4)
    rows = list(wb['single-hit'].iter_rows(values_only=True))[1:]
    wt   = next(r for r in rows if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
    return rows, wt[1], wt[2]

def load_pwms():
    wb, ws = openpyxl.load_workbook(MMC3), None
    ws = openpyxl.load_workbook(MMC3)['시트1']
    pwms = {}
    skip = {"Source: Cis-BP","Source: Weirauch14","Source: Jolma_367",
            "Source: Jolma_371","Source: Jolma_353","Source: Jolma_354",
            "Source: Badis09, Zhao and Stormo 11"}
    tf, cur = None, {"A":[],"C":[],"G":[],"T":[]}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[0] not in skip:
            if tf: pwms[tf] = cur
            tf, cur = row[0], {"A":[],"C":[],"G":[],"T":[]}
        if row[1] in "ACGT":
            cur[row[1]] = [v for v in row[2:] if v is not None]
    if tf: pwms[tf] = cur
    return pwms

def load_params_4tf(model_num=5):
    ws   = openpyxl.load_workbook(MMC2)['4TF_self']
    rows = list(ws.iter_rows(values_only=True))
    mrow = next(r for r in rows if r[0]=='TF' and r[1]=='Model num')
    mcol = {int(mrow[c]):c for c in range(2,10)}[model_num]
    p, tf = {}, None
    for row in rows:
        if row[0]=='TF': continue
        if row[0] is not None: tf = row[0]
        if row[1] in ['Rmax','Theta']: p[row[1]] = row[mcol]
        elif row[1] in ['coef','kmax','lambda','threshold'] and tf:
            p.setdefault(tf,{})[row[1]] = row[mcol]
    return p

def make_pwm_xml(d):
    lines = ['        <PWM type="PSSM">','          <base>A C G T</base>']
    for i in range(len(d['A'])):
        lines.append(f'          <position>{d["A"][i]:10.6f}; {d["C"][i]:10.6f}; {d["G"][i]:10.6f}; {d["T"][i]:10.6f}</position>')
    lines.append('        </PWM>')
    return "\n".join(lines)

def make_tf_xml(name, p, pwm):
    return f"""      <TF name="{name}" bsize="14" include="true">
        <kmax      value="{p['kmax']}"      lim_low="1e-06" lim_high="4.0"  anneal="false" move="Kmax"/>
        <threshold value="{p['threshold']}" lim_low="-5"   lim_high="15.0" anneal="false" move="Sites"/>
        <lambda    value="{p['lambda']}"    lim_low="0.5"  lim_high="5.0"  anneal="false" move="Lambda"/>
        <Coefficients>
          <coef value="{p['coef']}" lim_low="0.0001" lim_high="20.0" anneal="false" move="Promoter"/>
        </Coefficients>
{make_pwm_xml(pwm)}
      </TF>"""

def make_xml(params, pwms, rows, theta, tfs=['CREB1','CREM','ATF1','ATF7']):
    gene_lines = "\n".join([f'        <Gene name="{r[0]}" header="{r[0]}" left_bound="-235" right_bound="-38" TSS="-1" promoter="basic"/>' for r in rows if r[0]])
    gene_cols  = " ".join([f'{r[0]}="{r[2]:.4f}"' for r in rows if r[0] and r[2]])
    tfs_xml    = "\n".join([make_tf_xml(tf, params.get(tf,{}), pwms.get(tf, pwms['CREB1'])) for tf in tfs])
    return f"""<?xml version="1.0" encoding="utf-8"?>
<Root>
  <annealer_input init_T="100000" lambda="0.0001" init_loop="100000"/>
  <move interval="100" gain="3"/>
  <count_criterion freeze_crit="10" freeze_cnt="5"/>
  <mix adaptcoef="10"/>
  <lam tau="100" memLength_mean=".200" memLength_sd="100" criterion="10" freeze_cnt="5"/>
  <Mode>
    <Verbose value="1"/>
    <ScoreFunction value="sse"/>
    <GCcontent value="0.534"/>
    <NumThreads value="4"/>
    <SelfCompetition value="true"/>
    <Precision value="8"/>
    <Seed value="5"/>
  </Mode>
  <Input>
    <Distances>
      <Distance name="Quenching" distfunc="Trapezoid">
        <A value="100" lim_low="50" lim_high="100" anneal="false" move="Quenching"/>
        <B value="50"  lim_low="50" lim_high="50"  anneal="false" move="Quenching"/>
      </Distance>
    </Distances>
    <Promoters>
      <Promoter name="basic" function="Arrhenius2">
        <Q     value="1"   lim_low="1"   lim_high="1"   anneal="false" move="Promoter"/>
        <Rmax  value="200" lim_low="200" lim_high="200" anneal="false" move="Promoter"/>
        <Theta value="{theta}" lim_low="5" lim_high="20" anneal="false" move="Promoter"/>
      </Promoter>
    </Promoters>
    <TFs>
{tfs_xml}
    </TFs>
    <Interactions/>
    <Genes>
      <Source name="MRPA" file="sequences/MRPA_baseM.fa" type="fasta">
{gene_lines}
      </Source>
    </Genes>
    <ScaleFactors>
      <ScaleFactor name="default">
        <A value="1" lim_low="1" lim_high="1" anneal="false"/>
        <B value="0" lim_low="0" lim_high="0" anneal="false"/>
      </ScaleFactor>
    </ScaleFactors>
    <RateData row="ID" col="gene">
      <TableRow ID="theonlyone" {gene_cols}/>
    </RateData>
    <TFData row="ID" col="TF">
      <TableRow ID="theonlyone" {" ".join([f'{tf}="100"' for tf in tfs])}/>
    </TFData>
  </Input>
</Root>"""

def main():
    print("=== Figure 2C 생성 시작 ===")

    rows, wt_seq, wt_expr = load_single_hit()
    pwms   = load_pwms()
    params = load_params_4tf(model_num=5)
    theta  = params.get('Theta', 5.06)

    # 측정 Δactivity
    meas_dict = {r[0]: np.log2(float(r[2])/float(wt_expr))
                 for r in rows if r[0] and r[2] and re.search(r'pos_(\d+)_([ATGC])$', str(r[0]))}

    # FASTA 생성
    fasta_path = f"{FITS}/sequences/MRPA_baseM.fa"
    if not os.path.exists(fasta_path):
        os.makedirs(os.path.dirname(fasta_path), exist_ok=True)
        with open(fasta_path, 'w') as f:
            for name, seq, expr in rows:
                if seq:
                    f.write(f'>{name}\n')
                    for i in range(0, len(seq), 60):
                        f.write(seq[i:i+60] + '\n')
        print(f"FASTA 생성 완료: {fasta_path}")

    # XML 생성
    xml_path = f"{FITS}/fig2c_model5.xml"
    with open(xml_path, 'w') as f:
        f.write(make_xml(params, pwms, rows, theta))
    print(f"XML 생성: {xml_path}")

    # transcpp 실행
    rates_path = f"{xml_path}.rates"
    if not os.path.exists(rates_path):
        print("transcpp 실행 중... (1~2분 소요)")
        subprocess.run([TRANSCPP, xml_path], cwd=FITS)
    else:
        print(f"기존 rates 파일 사용")

    # R값 계산
    with open(rates_path) as f: lines = f.readlines()
    header = lines[0].strip().split()
    data   = lines[1].strip().split()
    pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
    wt_p   = pred['synCRE_Promega_0']
    pred_dict = {k: np.log2(pred[k]/wt_p) for k in meas_dict if k in pred and pred[k] > 0}

    vals_m = [meas_dict[k] for k in pred_dict]
    vals_p = [pred_dict[k] for k in pred_dict]
    r_val, _ = stats.pearsonr(vals_m, vals_p)
    rmse = np.sqrt(np.mean([(m-p)**2 for m,p in zip(vals_m, vals_p)]))
    print(f"Pearson R = {r_val:.3f}, RMSE = {rmse:.3f}")

    # 그래프
    bases = ['A','C','G','T']
    fig, axes = plt.subplots(4, 2, figsize=(16, 12))
    fig.suptitle(f'Figure 2C: 4TF Self-Competition Model (R={r_val:.3f})', fontsize=13, fontweight='bold')

    for ri, base in enumerate(bases):
        pos_m, pos_p = {}, {}
        for name, mv in meas_dict.items():
            m = re.search(r'pos_(\d+)_([ATGC])$', name)
            if m and m.group(2) == base:
                pos = int(m.group(1))
                pos_m[pos] = mv
                if name in pred_dict: pos_p[pos] = pred_dict[name]

        positions = sorted(pos_m.keys())
        for ci, (vals, color, label) in enumerate([
            ([pos_m.get(p,0) for p in positions], '#4472C4', f'→{base} measured'),
            ([pos_p.get(p,0) for p in positions], '#C00000', f'→{base} predicted'),
        ]):
            ax = axes[ri, ci]
            ax.bar(positions, vals, color=color, alpha=0.85, width=0.8)
            ax.axhline(0, color='black', linewidth=0.8)
            ax.set_xlim(-1, 87); ax.set_ylim(-2.5, 2.5)
            ax.set_title(label, fontsize=10)
            ax.set_facecolor('#F5F5F5')
            ax.grid(True, color='white', linewidth=0.8)
            if ci == 0: ax.set_ylabel('Δactivity (log2)', fontsize=9)
        if ri == 3:
            axes[ri,0].set_xlabel('Enhancer position', fontsize=10)
            axes[ri,1].set_xlabel('Enhancer position', fontsize=10)

    plt.tight_layout()
    out = f"{OUTPUT}/Figure2C.png"
    plt.savefig(out, dpi=150, bbox_inches='tight')
    print(f"✅ 저장: {out}")

if __name__ == '__main__':
    main()
