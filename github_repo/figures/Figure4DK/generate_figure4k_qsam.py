"""
Figure 4K: QSAM multi-hit prediction
선형 모델(Ridge regression)로 single-hit 학습 후 multi-hit 예측

사용법:
    python3 generate_figure4k_qsam.py

출력:
    Figure4DK/output/Figure4K.png
"""

import os
import numpy as np
import openpyxl
from scipy import stats
from sklearn.linear_model import Ridge
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

BASE   = os.path.expanduser("~/Desktop/SB_final/ScienceDirect_files_26May2026_04-41-20.049")
MMC4   = f"{BASE}/1-s2.0-S2589004223028249-mmc4.xlsx"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT, exist_ok=True)

BASE_IDX = {'A':0,'C':1,'G':2,'T':3}

def one_hot(seq, L=197):
    v = np.zeros(L * 4)
    for i, b in enumerate(seq[:L]):
        if b in BASE_IDX:
            v[i*4 + BASE_IDX[b]] = 1
    return v

def main():
    print("=== Figure 4K: QSAM 생성 시작 ===")

    wb     = openpyxl.load_workbook(MMC4)
    rows_s = list(wb['single-hit'].iter_rows(values_only=True))[1:]
    rows_m = list(wb['multi-hit'].iter_rows(values_only=True))[1:]
    wt_s   = next(r[2] for r in rows_s if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
    wt_m   = next(r[2] for r in rows_m if 'WT' in str(r[0]))
    SEQ_L  = len(next(r[1] for r in rows_s if r[1]))

    # 학습 (single-hit)
    train_X = np.array([one_hot(r[1], SEQ_L) for r in rows_s if r[1] and r[2]])
    train_y = np.array([np.log2(float(r[2])/float(wt_s)) for r in rows_s if r[1] and r[2]])
    model   = Ridge(alpha=0.01)
    model.fit(train_X, train_y)
    r_train, _ = stats.pearsonr(train_y, model.predict(train_X))
    print(f"Training R = {r_train:.4f}")

    # 예측 (multi-hit)
    meas, pred = [], []
    for name, seq, expr in rows_m:
        if not seq or not expr or 'WT' in str(name): continue
        if len(seq) != SEQ_L: continue
        meas.append(np.log2(float(expr)/float(wt_m)))
        pred.append(model.predict(one_hot(seq, SEQ_L).reshape(1,-1))[0])

    meas, pred = np.array(meas), np.array(pred)
    r_val, _   = stats.pearsonr(meas, pred)
    rmse       = np.sqrt(np.mean((meas-pred)**2))
    print(f"Multi-hit R = {r_val:.3f}, RMSE = {rmse:.3f}  (논문 목표: R=0.87)")

    # 그래프
    fig, ax = plt.subplots(figsize=(6, 6))
    ax.set_facecolor('#F0F0F0')
    ax.grid(True, color='white', linewidth=1.2, zorder=0)
    ax.scatter(meas, pred, alpha=0.4, s=15, color='#4472C4', zorder=3)

    # inset (training)
    ax_in = ax.inset_axes([0.05, 0.6, 0.35, 0.35])
    ax_in.scatter(train_y, model.predict(train_X), alpha=0.3, s=5, color='#4472C4')
    ax_in.text(0.05, 0.95, f'R={r_train:.2f}', transform=ax_in.transAxes, va='top', fontsize=8)
    ax_in.tick_params(labelsize=5)

    ax.set_xlim(-4, 4); ax.set_ylim(-4, 4)
    ax.set_xlabel('MPRA multi-hit data', fontsize=12)
    ax.set_ylabel('QSAM prediction', fontsize=12)
    ax.set_title('Figure 4K: QSAM Multi-hit Prediction', fontsize=11, fontweight='bold')
    ax.text(0.05, 0.97, f'R={r_val:.2f}', transform=ax.transAxes, va='top', fontsize=13, fontweight='bold')
    ax.tick_params(axis='both', which='both', length=0)
    for spine in ax.spines.values(): spine.set_visible(False)

    plt.tight_layout()
    out = f"{OUTPUT}/Figure4K.png"
    plt.savefig(out, dpi=150, bbox_inches='tight')
    print(f"✅ 저장: {out}")

if __name__ == '__main__':
    main()
