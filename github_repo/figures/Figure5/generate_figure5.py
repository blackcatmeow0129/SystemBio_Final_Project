"""
Figure 5A/E/G 재현 스크립트
- 5A: →A 치환 Δactivity bar plot
- 5E: TF activation coefficient
- 5G: →T 치환 Δactivity bar plot

사용법:
    cd ~/cre_reproduction/neoParSA/tests/transcpp/fits
    python3 /path/to/generate_figure5.py

출력:
    Figure5/output/Figure5AG.png
    Figure5/output/Figure5E.png
"""

import os, re
import numpy as np
import openpyxl
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

BASE   = os.path.expanduser("~/Desktop/SB_final/ScienceDirect_files_26May2026_04-41-20.049")
FITS   = os.path.expanduser("~/cre_reproduction/neoParSA/tests/transcpp/fits")
MMC2   = f"{BASE}/1-s2.0-S2589004223028249-mmc2.xlsx"
MMC4   = f"{BASE}/1-s2.0-S2589004223028249-mmc4.xlsx"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT, exist_ok=True)

# CRE 위치 (87bp 기준, 0-indexed)
CRE_REGIONS = {
    'CRE1':   (8,  19,  '#FFB3B3'),
    'CRE2':   (35, 44,  '#FFB3B3'),
    'CRE3':   (44, 52,  '#FFB3B3'),
    'cryptic':(60, 68,  '#B3D9FF'),
    'CRE4':   (67, 79,  '#FFB3B3'),
}

def load_single_hit():
    wb   = openpyxl.load_workbook(MMC4)
    rows = list(wb['single-hit'].iter_rows(values_only=True))[1:]
    wt   = next(r[2] for r in rows if 'Promega_0' in str(r[0]) and 'scanmut' not in str(r[0]))
    return rows, wt

def load_rates(rates_path):
    with open(rates_path) as f: lines = f.readlines()
    header = lines[0].strip().split()
    data   = lines[1].strip().split()
    pred   = {header[i+1]: float(data[i+1]) for i in range(len(header)-1)}
    wt_p   = pred.get('synCRE_Promega_0', 1.0)
    return pred, wt_p

def load_coef():
    ws    = openpyxl.load_workbook(MMC2)['4TF_self']
    rows  = list(ws.iter_rows(values_only=True))
    mrow  = next(r for r in rows if r[0]=='TF' and r[1]=='Model num')
    mcol  = {int(mrow[c]):c for c in range(2,10)}[5]
    coefs = {}
    tf    = None
    for row in rows:
        if row[0]=='TF': continue
        if row[0] is not None: tf = row[0]
        if row[1]=='coef' and tf: coefs[tf] = row[mcol]
    return coefs

def plot_5ag(rows, wt_expr, pred, wt_p, output_path):
    """Figure 5A (→A) and 5G (→T) bar plots"""
    fig, axes = plt.subplots(2, 2, figsize=(16, 8))
    fig.suptitle('Figure 5A/G: Δactivity for →A and →T substitutions', fontsize=12, fontweight='bold')

    for row_idx, (base, row_label) in enumerate([('A','A'),('T','G')]):
        pos_meas, pos_pred = {}, {}
        for r in rows:
            name, seq, expr = r
            m = re.search(r'pos_(\d+)_([ATGC])$', str(name))
            if m and m.group(2) == base and expr:
                pos = int(m.group(1))
                pos_meas[pos] = np.log2(float(expr)/float(wt_expr))
                if name in pred and pred[name] > 0:
                    pos_pred[pos] = np.log2(pred[name]/wt_p)

        positions = sorted(pos_meas.keys())

        for ci, (vals_dict, color, label) in enumerate([
            (pos_meas, '#4472C4', f'→{base} MPRA'),
            (pos_pred, '#C00000', f'→{base} Model'),
        ]):
            ax = axes[row_idx, ci]
            vals = [vals_dict.get(p, 0) for p in positions]
            ax.bar(positions, vals, color=color, alpha=0.8, width=0.8, zorder=3)
            ax.axhline(0, color='black', linewidth=0.8)

            # CRE 영역 표시
            for name_r, (start, end, color_r) in CRE_REGIONS.items():
                ax.axvspan(start, end, alpha=0.2, color=color_r, zorder=1)

            ax.set_xlim(-1, 87)
            ax.set_ylim(-2.5, 2.5)
            ax.set_title(label, fontsize=10)
            ax.set_facecolor('#F8F8F8')
            ax.grid(True, color='white', linewidth=0.8, zorder=0)
            if ci == 0: ax.set_ylabel('Δactivity (log2)', fontsize=9)
            if row_idx == 1: ax.set_xlabel('Enhancer position', fontsize=9)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    print(f"✅ 저장: {output_path}")

def plot_5e(coefs, output_path):
    """Figure 5E: TF activation coefficients"""
    tfs  = ['CREB1','CREM','ATF1','ATF7']
    vals = [np.log10(coefs.get(tf, 1.0)) for tf in tfs]
    colors = ['#FF7F00','#E41A1C','#377EB8','#4DAF4A']

    fig, ax = plt.subplots(figsize=(5, 5))
    bars = ax.bar(tfs, vals, color=colors, alpha=0.85, width=0.6)
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_ylabel('log10(coef)', fontsize=12)
    ax.set_title('Figure 5E: TF Activation Coefficients', fontsize=11, fontweight='bold')
    ax.set_facecolor('#F8F8F8')
    ax.grid(True, axis='y', color='white', linewidth=0.8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    for bar, val in zip(bars, vals):
        ax.text(bar.get_x()+bar.get_width()/2, val+0.05, f'{10**val:.2f}',
                ha='center', va='bottom', fontsize=9)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    print(f"✅ 저장: {output_path}")

def main():
    print("=== Figure 5A/E/G 생성 시작 ===")

    rows, wt_expr = load_single_hit()
    coefs = load_coef()
    print(f"4TF coef: { {tf: f'{v:.3f}' for tf, v in coefs.items()} }")

    # rates 파일 (Figure 2C에서 생성된 것 재활용)
    rates_path = f"{FITS}/fig2c_model5.xml.rates"
    if os.path.exists(rates_path):
        pred, wt_p = load_rates(rates_path)
        plot_5ag(rows, wt_expr, pred, wt_p, f"{OUTPUT}/Figure5AG.png")
    else:
        print(f"⚠️  rates 파일 없음: {rates_path}")
        print("    먼저 generate_figure2c.py를 실행하세요")

    plot_5e(coefs, f"{OUTPUT}/Figure5E.png")
    print("=== Figure 5 완료 ===")

if __name__ == '__main__':
    main()
